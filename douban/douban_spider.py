# -*- coding: utf-8 -*-
"""
    @Author  : shenxuexin
    @Time    : 2020/12/29 21:20
    @File    : douban_spider.py
    @software:PyCharm
"""
import gevent
from gevent import monkey, pool
monkey.patch_all()

import sys
import os
import time
import re
import requests
from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import datetime
from config import *
from loguru import logger
from bs4 import BeautifulSoup


class DoubanSpider(object):
    def __init__(self, types=None):
        self.types = types if types else []
        self.session = requests.Session()
        self.session.headers.update({'User-Agent': USER_AGENT})
        self.image_download_path = IMG_DOWNLOAD_PATH
        self.file_save_path = FILE_SAVE_PATH
        self.excel_save_path = os.path.join(self.file_save_path, 'book_info.xlsx')

        if not os.path.exists(self.image_download_path):
            os.makedirs(self.image_download_path)

        if not os.path.exists(self.file_save_path):
            os.makedirs(self.file_save_path)

    def safe_get(self, url, params=None):
        try:
            response = self.session.get(url, params=params)
            if response.status_code == 403:
                logger.error(f'Anti scraper. exit.')
                sys.exit(0)
            elif response.status_code == 404:
                logger.error(f'404 Not found.')
        except Exception as e:
            logger.error(f'Request failed. | Exception: {e}')
            return

        time.sleep(SLEEP_PER_REQUEST)
        return response

    def get_book_category(self):
        book_url = BOOK_URL
        cate_response = self.safe_get(book_url)
        if not cate_response:
            return

        cate_bs = BeautifulSoup(cate_response.text, 'html.parser')
        cate_list_ul = cate_bs.find('ul', {'class': 'type-list'})
        cate_li_list = cate_list_ul.find_all('li')

        cate_info_list = []
        for cate_li in cate_li_list:
            cate_url = cate_li.a['href']
            cate_name = cate_li.get_text()

            cate_info_list.append((cate_name, BASIC_SERVER_URL + cate_url))

        logger.info(f'Get book category list successful. | list: {[cate_info[0] for cate_info in cate_info_list]}')
        return cate_info_list

    def get_keyword(self, cate_url):
        resp = self.safe_get(cate_url)

        if resp:
            keyword = ''
            try:
                match = re.search(f'var\s*TYPE\s*=\s*\'(\w+)\'', resp.text)
                if match:
                    keyword = match.group(1)
            except Exception as e:
                logger.error(f'Get keyword failed. | Exception: {e}')
                return

            return keyword

    def get_books_info(self, cate_info, book_count_dict):
        cate_name_cn, cate_url = cate_info
        logger.info(f'Start getting book info. | category: {cate_name_cn}')

        keyword = self.get_keyword(cate_url)
        if not keyword:
            return

        book_info_url = BOOK_INFO_URL.format(keyword)
        self.session.headers.update({'Referer': cate_url})

        num = 0
        per_page = 18
        total = 1
        book_info_list = []
        while num < total:
            args = {'start': num, 'count': per_page}
            resp = self.safe_get(book_info_url, params=args)

            try:
                book_page_info = resp.json()
            except Exception as e:
                logger.error(f'Get book info failed. | Exception: {e}')
                break

            book_item_list = book_page_info.get('subject_collection_items', [])
            if not book_item_list:
                logger.warning(f'book_page_info: {book_page_info}')
            for book_item in book_item_list:
                book_id = book_item.get('id')
                if not book_id:
                    continue

                title = book_item.get('title', '未知标题')
                author = book_item.get('author', ['未知作者'])
                tags = book_item.get('tags', '')

                cover_url = book_item['cover']['url']
                try:
                    value = book_item['rating']['value']
                except:
                    logger.warning(f'No value. | book: {title}')
                    value = -1

                book_info = {'id': book_id,
                             'title': title,
                             'author': ','.join(author),
                             'tags': ','.join(tags),
                             'value': value}

                # 下载封面
                # img_path = os.path.join(self.image_download_path, f'{book_id}.jpg')
                # self.image_download(cover_url, path=img_path)

                book_info_list.append(book_info)

            total = book_page_info.get('total', 0)
            num += per_page

        book_count_dict.update({cate_name_cn: len(book_info_list)})
        book_info_dict = {'category': cate_name_cn, 'book_info_list': book_info_list}
        logger.info(f'Get book info successful. | category: {cate_name_cn} | count: {len(book_info_list)}')
        return book_info_dict

    def image_download(self, url, path):
        img_resp = self.safe_get(url)
        try:
            with open(path, 'wb') as f:
                f.write(img_resp.content)
        except Exception as e:
            logger.error(f'Download image failed. | url: {url} | Exception: {e}')

        logger.info(f'Download image successful. | path: {path}')

    def book_info_save(self, book_page_info):
        if not os.path.exists(self.excel_save_path):
            workbook = Workbook()
            _ = workbook.active
        else:
            workbook = load_workbook(self.excel_save_path)

        category = book_page_info.get('category', '')
        sheet_name_list = workbook.sheetnames
        if category in sheet_name_list:
            logger.info(f'sheet already exists. | category: {category}')
            return
        else:
            book_sheet = workbook.create_sheet(category)

        table_head = ['id', '标题', '作者', '标签', '评分']
        book_sheet.append(table_head)

        book_info_list = book_page_info.get('book_info_list', [])
        for book_info in book_info_list:
            book_id = book_info.get('id')
            title = book_info.get('title')
            author = book_info.get('author')
            tags = book_info.get('tags')
            value = book_info.get('value')

            cell_info_list = [book_id, title, author, tags, value]
            book_sheet.append(cell_info_list)

        workbook.save(self.excel_save_path)
        logger.info(f'Save excel successful. | path: {self.excel_save_path}')

    def scrape_per_category(self, book_cate_info, book_count_dict):
        books_info = self.get_books_info(book_cate_info, book_count_dict)
        if not books_info:
            return

        self.book_info_save(books_info)
        time.sleep(SLEEP_PER_CATEGORY)

    def main(self):
        # 抓取图书
        time_start = datetime.now().replace(microsecond=0)
        logger.info(f'Star scraping. | time: {time_start}')
        if 'book' in self.types:
            book_cate_list = self.get_book_category()
            if not book_cate_list:
                return

            book_count_dict = {book_cate_info[0]: 0 for book_cate_info in book_cate_list}
            cate_pool = pool.Pool(BATCH_PER_GEVENT)
            for book_cate_info in book_cate_list:
                cate_gl = gevent.spawn(self.scrape_per_category, book_cate_info, book_count_dict)
                cate_pool.add(cate_gl)
            cate_pool.join()

            time_span = (datetime.now() - time_start).seconds
            logger.info(f'Scrape books done. | time_span: {time_span}s | count: {book_count_dict}')


if __name__ == '__main__':
    douban_spider = DoubanSpider(types=['book'])
    douban_spider.main()
