# -*- coding: utf-8 -*-
"""
    @Author  : shenxuexin
    @Time    : 2021/1/3 18:04
    @File    : qiubai_spider.py
    @software:PyCharm
"""
import gevent
from gevent import monkey, pool
monkey.patch_all()

import requests
import time
import sys
import os

from openpyxl import Workbook, load_workbook
from datetime import datetime
from lxml import etree
from loguru import logger
from config import *


class QiubaiSpider(object):
    def __init__(self, mode):
        self.session = requests.Session()
        self.session.headers.update({'User-Agent': USER_AGENT})
        self.mode = mode
        self.file_save_path = FILE_SAVE_PATH
        self.excel_save_path = os.path.join(self.file_save_path, 'qiubai.xlsx')
        self.article_id_list = []

        if self.mode not in MODE_LIST:
            logger.error(f'Input mode not be supported. | mode: {mode}')
            sys.exit(0)

        if not os.path.exists(self.file_save_path):
            os.makedirs(self.file_save_path)

    def safe_get(self, url, params=None):
        try:
            resp = self.session.get(url, params=params)

            if resp.status_code == 403:
                logger.warning(f'Anti scraper.')
                time.sleep(SLEEP_PER_ANTI_SCRAPER)
            elif resp.status_code == 404:
                logger.warning(f'404 not found.')
        except Exception as e:
            logger.error(f'Request failed. | url: {url} | Exception: {e}')
            return

        time.sleep(SLEEP_PER_REQUEST)
        return resp

    def get_url_list(self):
        dest_url = MODE_URL_DICT.get(self.mode)
        url_list = [BASIC_SERVER_URL + dest_url.format(i) for i in range(1, 14)]
        logger.info(f'Get url list successfully. | url_list: {url_list}')
        return url_list

    def content_extract(self, resp):
        html = etree.HTML(resp.text)
        div_list = html.xpath('//div[@class="col1 old-style-col1"]/div')
        content_list = []
        for div in div_list:
            article_id = div.xpath('.//a[@class="contentHerf"]/@href')[0].split('/')[-1] if div.xpath('.//a[@class="contentHerf"]/@href') else None
            text = div.xpath('.//div[@class="content"]/span/text()') if div.xpath('.//div[@class="content"]/span/text()') else None
            image = 'http:' + div.xpath('.//div[@class="thumb"]/a/img/@src')[0] if div.xpath('.//div[@class="thumb"]/a/img/@src') else None
            video = 'http:' + div.xpath('.//video/source/@src')[0] if div.xpath('.//video/source/@src') else None
            author_img = 'http:' + div.xpath('//div[contains(@class,"author")]//img/@src')[0].split('?')[0] if div.xpath('.//div[contains(@class,"author")]//img/@src') else None
            author_name = div.xpath('.//div[contains(@class,"author")]//h2/text()')[0].replace('\n', '') if div.xpath('.//div[contains(@class, "author")]//h2/text()') else None
            author_gender = div.xpath('.//div[contains(@class, "articleGender")]/@class')[0].split()[-1][:-4] if div.xpath('.//div[contains(@class, "articleGender")]/@class') else None
            author_age = div.xpath('.//div[contains(@class, "articleGender")]/text()')[0] if div.xpath('.//div[contains(@class, "articleGender")]/text()') else None
            stats_vote = div.xpath('//span[@class="stats-vote"]/i/text()')[0] if div.xpath('//span[@class="stats-vote"]/i/text()') else None

            if text:
                text = ''.join(text).replace('\n', '')

            if article_id and article_id not in self.article_id_list:
                self.article_id_list.append(article_id)
                content_dict = {'id': article_id,
                                'text': text,
                                'image': image,
                                'video': video,
                                'author_img': author_img,
                                'author_name': author_name,
                                'author_gender': author_gender,
                                'author_age': author_age,
                                'stats_vote': stats_vote}
                content_list.append(content_dict)

        logger.info(f'Content extract successful.')
        return content_list

    def content_save(self, content_list):
        if not os.path.exists(self.excel_save_path):
            workbook = Workbook()
            _ = workbook.active
        else:
            workbook = load_workbook(self.excel_save_path)

        sheet_name_list = workbook.sheetnames
        if self.mode in sheet_name_list:
            article_sheet = workbook[self.mode]
        else:
            article_sheet = workbook.create_sheet(self.mode)
            table_head = ['id', 'text', 'image', 'video', 'avatar', 'username', 'gender', 'age', 'vote']
            article_sheet.append(table_head)

        for article in content_list:
            cell_info_list = [u'无' if i is None else i for i in article.values()]
            article_sheet.append(cell_info_list)

        workbook.save(self.excel_save_path)
        logger.info(f'Save excel successful. | path: {self.excel_save_path}')

    def scrape(self, url):
        resp = self.safe_get(url)
        content = self.content_extract(resp)
        self.content_save(content)

    def main(self):
        time_begin = datetime.now().replace(microsecond=0)
        logger.info(f'Start scraping. | mode: {self.mode} | time: {time_begin}')
        url_list = self.get_url_list()
        scrape_pool = pool.Pool(PER_BATCH_FOR_GEVENT)
        for url in url_list:
            scrape_gl = gevent.spawn(self.scrape, url)
            scrape_pool.add(scrape_gl)
        scrape_pool.join()

        logger.info(f'Scrape done. | article: {len(self.article_id_list)} | time_span: {(datetime.now()-time_begin).seconds}s')


if __name__ == '__main__':
    mode_list = ['text', 'image', 'video']
    for mode in mode_list:
        qiubai_spider = QiubaiSpider(mode=mode)
        qiubai_spider.main()